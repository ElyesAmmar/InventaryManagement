from model import Products
from openpyxl import Workbook, load_workbook


def create_product_table():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Products"
    headers = ("product_id", "name", "price", "stock")
    sheet.append(headers)
    wb.save("products.xlsx")
    print("Excel file created successfully!")


def add_products():

    # get data from user
    product_id = input("enter id product:")
    name = input("enter name product:")
    price = float(input("enter price product:"))
    stock = int(input("enter stock product:"))
    products = Products(product_id, name, price, stock).product_data()
    print(products)

    # add data to xl file products
    filename = "products.xlsx"
    wb = load_workbook(filename)
    sheet = wb["Products"]
    row = list(products.values())
    sheet.append(row)
    wb.save("products.xlsx")
    print("Product created successfully!")


add_products()
# create_product_table()
